from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from Excel2RpyScript.handler.converter import Converter
from Excel2RpyScript.handler.parser import Parser
from Excel2RpyScript.handler.writer import RpyFileWriter

def script_preprocess(input_filepath, output_filepath):
    # 加载工作簿
    workbook = load_workbook(input_filepath)

    # 获取所需列的索引
    col_a_index = column_index_from_string("A")
    col_v_index = column_index_from_string("V")
    col_ab_index = column_index_from_string("AB")
    col_z_index = column_index_from_string("Z")
    col_x_index = column_index_from_string("X")
    # 遍历每个工作表
    for sheet in workbook.worksheets:
        last_non_empty_identifier = None  # 用于跟踪V列的最近非空的行标识符
        last_non_empty_value = None  # 用于跟踪AB列的最近非空的A列值
        value_counts = {}  # 用于跟踪每个值的出现次数

        # 从第8行开始遍历每一行
        for row in sheet.iter_rows(min_row=8):
            cell_a = row[col_a_index - 1]  # 获取A列的单元格
            cell_v = row[col_v_index - 1]  # 获取V列的单元格
            cell_ab = row[col_ab_index - 1]  # 获取AB列的单元格
            cell_z = row[col_z_index - 1]  # 获取Z列的单元格
            cell_x = row[col_x_index - 1]  # 获取X列的单元格

            # 检查A列单元格是否为空
            if cell_a.value:
                last_non_empty_value = cell_a.value
                last_non_empty_identifier = f"{sheet.title.lower()}_{cell_a.row}"

                # 根据最近非空的行标识符设置V列的值
                cell_v.value = last_non_empty_identifier
                # 如果Z列为空，则默认添加”溶解“作为转场
                if cell_z.value is None:
                    cell_z.value = "溶解"
                # 如果X列为空，则默认添加”adv“作为模式
                if cell_x.value is None:
                    cell_x.value = "adv"

            # 更新该值的出现次数并设置AB列的值
            # 这是每个角色的配音从1开始编号
            # value_counts[last_non_empty_value] = value_counts.get(last_non_empty_value, 0) + 1
            # 这是所有角色的配音放在一起，从1开始编号
            all = "all"
            value_counts[all] = value_counts.get(all, 0) + 1
            cell_ab.value = f"{sheet.title}_{last_non_empty_value}_{value_counts[all]-1}.wav"

    # 保存到新的文件
    workbook.save(output_filepath)

def script_process(input_filepath, output_filepath):
    # 1. 解析Excel文件
    file_path = str(input_filepath)
    parser = Parser(file_path)
    parsed_sheets = parser.get_parsed_sheets()

    # 2. 将解析的数据转换为Rpy元素
    converter = Converter(parser)
    rpy_elements = converter.generate_rpy_elements()

    # 3. 将Rpy元素写入到.rpy文件中
    output_dir = str(output_filepath)
    for res in rpy_elements:
        RpyFileWriter.write_file(output_dir, res, converter.role_name_mapping, converter.side_characters)
# # 使用函数
# input_filepath = "工作簿1.xlsx"
# output_filepath = "your_output_file_path.xlsx"
# script_preprocess(input_filepath, output_filepath)
