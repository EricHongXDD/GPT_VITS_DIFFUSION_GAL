# -*- coding: utf-8 -*-
from datetime import datetime
import os

# Form implementation generated from reading ui file 'script_window.ui'
#
# Created by: PyQt5 UI code generator 5.15.9
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QStringListModel, QThread, pyqtSignal
from PyQt5.QtWidgets import QFileDialog, QHeaderView, QTableWidgetItem
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from script_preprocess import script_preprocess

# 脚本生成线程类
class ScriptThread(QThread):
    # 更新日志的信号量
    string_signal = pyqtSignal(str)

    def __init__(self, solve_script_path):
        super(ScriptThread, self).__init__()
        self.solve_script_path = solve_script_path
        self.running = True

    def run(self):

        self.string_signal.emit('脚本生成完成')

class Ui_Script(object):
    def setupUi(self, Script):
        Script.setObjectName("Script")
        Script.resize(1711, 1041)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        Script.setPalette(palette)
        # self.centralwidget = QtWidgets.QWidget(Script)
        # self.centralwidget.setObjectName("centralwidget")
        self.DrawListView = QtWidgets.QListView(Script)
        self.DrawListView.setGeometry(QtCore.QRect(1350, 11, 341, 241))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.DrawListView.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.DrawListView.setFont(font)
        self.DrawListView.setStyleSheet("background-color: rgb(24, 24, 24); color: rgb(255, 255, 255);")
        self.DrawListView.setObjectName("DrawListView")
        self.layoutWidget_2 = QtWidgets.QWidget(Script)
        self.layoutWidget_2.setGeometry(QtCore.QRect(20, 10, 1331, 47))
        self.layoutWidget_2.setObjectName("layoutWidget_2")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.layoutWidget_2)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.ImportConfigurationButton = QtWidgets.QPushButton(self.layoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.ImportConfigurationButton.sizePolicy().hasHeightForWidth())
        self.ImportConfigurationButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.ImportConfigurationButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.ImportConfigurationButton.setFont(font)
        self.ImportConfigurationButton.setStyleSheet("background-color: rgb(0, 104, 0);")
        self.ImportConfigurationButton.setObjectName("ImportConfigurationButton")
        self.horizontalLayout.addWidget(self.ImportConfigurationButton)
        self.SuggestScribesButton = QtWidgets.QPushButton(self.layoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.SuggestScribesButton.sizePolicy().hasHeightForWidth())
        self.SuggestScribesButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.SuggestScribesButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.SuggestScribesButton.setFont(font)
        self.SuggestScribesButton.setStyleSheet("background-color: rgb(0, 104, 0);")
        self.SuggestScribesButton.setObjectName("SuggestScribesButton")
        self.horizontalLayout.addWidget(self.SuggestScribesButton)
        self.StartScriptButton = QtWidgets.QPushButton(self.layoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.StartScriptButton.sizePolicy().hasHeightForWidth())
        self.StartScriptButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.StartScriptButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.StartScriptButton.setFont(font)
        self.StartScriptButton.setStyleSheet("background-color: rgb(0, 104, 0);")
        self.StartScriptButton.setObjectName("StartScriptButton")
        self.horizontalLayout.addWidget(self.StartScriptButton)
        self.PauseButton = QtWidgets.QPushButton(self.layoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.PauseButton.sizePolicy().hasHeightForWidth())
        self.PauseButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.PauseButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.PauseButton.setFont(font)
        self.PauseButton.setStyleSheet("background-color: rgb(222, 0, 0);")
        self.PauseButton.setObjectName("PauseButton")
        self.horizontalLayout.addWidget(self.PauseButton)
        self.ViewDrawFileButton = QtWidgets.QPushButton(self.layoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.ViewDrawFileButton.sizePolicy().hasHeightForWidth())
        self.ViewDrawFileButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.ViewDrawFileButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.ViewDrawFileButton.setFont(font)
        self.ViewDrawFileButton.setStyleSheet("background-color: rgb(0, 104, 0);")
        self.ViewDrawFileButton.setObjectName("ViewDrawFileButton")
        self.horizontalLayout.addWidget(self.ViewDrawFileButton)
        self.tableWidget = QtWidgets.QTableWidget(Script)
        self.tableWidget.setGeometry(QtCore.QRect(40, 270, 1651, 731))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tableWidget.sizePolicy().hasHeightForWidth())
        self.tableWidget.setSizePolicy(sizePolicy)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.label_2 = QtWidgets.QLabel(Script)
        self.label_2.setGeometry(QtCore.QRect(41, 159, 108, 37))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.SuggestKeywordsProgressBar = QtWidgets.QProgressBar(Script)
        self.SuggestKeywordsProgressBar.setGeometry(QtCore.QRect(41, 104, 1271, 49))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.SuggestKeywordsProgressBar.setFont(font)
        self.tableWidget.setFont(font)
        self.SuggestKeywordsProgressBar.setStyleSheet("QProgressBar {\n"
"   border: 2px solid grey;\n"
"   border-radius: 5px;\n"
"   background-color: #FFFFFF;\n"
"}\n"
"\n"
"QProgressBar::chunk {\n"
"   background-color: #0000FF;\n"
"   width: 20px;\n"
"}\n"
"\n"
"QProgressBar {\n"
"   border: 2px solid grey;\n"
"   border-radius: 5px;\n"
"   text-align: center;\n"
"}")
        self.SuggestKeywordsProgressBar.setProperty("value", 0)
        self.SuggestKeywordsProgressBar.setObjectName("SuggestKeywordsProgressBar")
        self.label = QtWidgets.QLabel(Script)
        self.label.setGeometry(QtCore.QRect(41, 61, 135, 37))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.SheetComboBox = QtWidgets.QComboBox(Script)
        self.SheetComboBox.setGeometry(QtCore.QRect(41, 202, 161, 43))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.SheetComboBox.setFont(font)
        self.tableWidget.setFont(font)

        self.SheetComboBox.setStyleSheet("border: 2px solid grey;\n"
"border-radius: 5px;")
        self.SheetComboBox.setObjectName("SheetComboBox")
        # Script.setCentralWidget(self.centralwidget)
        # self.menubar = QtWidgets.QMenuBar(Script)
        # self.menubar.setGeometry(QtCore.QRect(0, 0, 1711, 23))
        # self.menubar.setObjectName("menubar")
        # Script.setMenuBar(self.menubar)
        # self.statusbar = QtWidgets.QStatusBar(Script)
        # self.statusbar.setObjectName("statusbar")
        # Script.setStatusBar(self.statusbar)

        self.retranslateUi(Script)
        QtCore.QMetaObject.connectSlotsByName(Script)

        #设置tablewidget为不可编辑。其中，QtWidgets.QAbstractItemView.NoEditTriggers是一个枚举值，用于表示没有编辑触发器，这意味着用户不能编辑QTableWidget中的项。
        self.tableWidget.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)


        # 连接SheetComboBox的信号到changeSheet函数
        self.SheetComboBox.currentIndexChanged[str].connect(self.changeSheet)

        self.log_data = []
        self.role_file = "场景人物生成.xlsx"
        # 绑定button到槽函数
        self.ImportConfigurationButton.clicked.connect(self.open_file)
        self.SuggestScribesButton.clicked.connect(self.extract_data_from_excel)
        self.ViewDrawFileButton.clicked.connect(self.viewfile)
        self.StartScriptButton.clicked.connect(self.script)

    def retranslateUi(self, Script):
        _translate = QtCore.QCoreApplication.translate
        Script.setWindowTitle(_translate("Script", "MainWindow"))
        self.ImportConfigurationButton.setText(_translate("Script", "从Excel文件导入配置"))
        self.SuggestScribesButton.setText(_translate("Script", "推导场景人物文本描述"))
        self.StartScriptButton.setText(_translate("Script", "开始生成脚本"))
        self.PauseButton.setText(_translate("Script", "一键暂停"))
        self.ViewDrawFileButton.setText(_translate("Script", "查看文件"))
        self.label_2.setText(_translate("Script", "选择表单"))
        self.label.setText(_translate("Script", "推导词进度"))

    def set_listView_content(self):
        # 使用QStringListModel来设置listView的数据
        model = QStringListModel()
        model.setStringList(self.log_data)
        self.DrawListView.setModel(model)

    def add_string_to_listView(self, new_string):
        # 获取当前时间
        current_time = datetime.now()
        # 将时间转换为字符串格式
        time_string = current_time.strftime('%m-%d %H:%M:%S ')
        new_string = time_string + new_string
        self.log_data.append(new_string)
        # 更新 listView 的内容
        self.set_listView_content()

    def open_file(self):
        self.SheetComboBox.currentIndexChanged[str].disconnect(self.changeSheet)
        self.open_filepath, _ = QFileDialog.getOpenFileName(None, '选择Excel配置文件(路径尽量不要有中文)', '',
                                                   '(*.xlsx *.xls)')
        try:
            # 替换为单反斜杠
            self.open_filepath = str(self.open_filepath).replace("/", "\\").replace(":", ":")
            self.open_filename = os.path.basename(self.open_filepath)
            self.add_string_to_listView("打开文件路径" + self.open_filepath)
            print(self.open_filepath)
            # 加载Excel文件中的所有sheet名到ComboBox中
            self.workbook = load_workbook(self.open_filepath)
            self.SheetComboBox.clear()
            self.SheetComboBox.addItems(self.workbook.sheetnames)

            # 默认加载第一个sheet的数据
            self.loadExcelData(self.workbook.sheetnames[0])
            self.SheetComboBox.currentIndexChanged[str].connect(self.changeSheet)
        except Exception as e:
            self.add_string_to_listView("取消打开文件")
            print("取消打开文件", e)
            self.SheetComboBox.currentIndexChanged[str].connect(self.changeSheet)

    def changeSheet(self, sheet_name):
            """根据所选择的sheet名称加载数据"""
            self.loadExcelData(sheet_name)

    def loadExcelData(self, sheet_name):
        # 清空tableWidget的内容
        self.tableWidget.setRowCount(0)
        self.tableWidget.setColumnCount(0)

        self.active_sheet = self.workbook[sheet_name]

        # 获取要跳过的列的索引
        start_skip_col = column_index_from_string("C") - 1  # 转换为0-based索引
        end_skip_col = column_index_from_string("R") - 1

        # 设置header
        headers = ["角色", "话", "音乐", "立绘", "换页", "背景", "备注", "模式", "音效", "转场", "特殊效果", "语音",
                   "分支", "头像"]
        self.tableWidget.setColumnCount(len(headers))
        self.tableWidget.setHorizontalHeaderLabels(headers)

        # 从第8行开始读取数据
        for row_idx, row in enumerate(self.active_sheet.iter_rows(min_row=8)):
            self.tableWidget.insertRow(row_idx)

            # 过滤出除C-R之外的列
            filtered_row = row[:start_skip_col] + row[end_skip_col + 1:]

            for col_idx, cell in enumerate(filtered_row):
                cell_value = "" if cell.value is None else str(cell.value)
                self.tableWidget.setItem(row_idx, col_idx, QTableWidgetItem(cell_value))
            # 设置列宽
            self.tableWidget.setColumnWidth(1,600)
            self.tableWidget.setColumnWidth(5,150)
            self.tableWidget.setColumnWidth(11,220)
            # 设置行高
            for rowIndex in range(self.tableWidget.rowCount()):
                self.tableWidget.setRowHeight(rowIndex, 100)
            # 列自适应，(水平方向占满窗体)
            # self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def extract_data_from_excel(self):
        try:
            workbook = load_workbook(self.open_filepath)

            data = []
            for sheet in workbook.worksheets:
                # 使用列表存储提取的数据，元素是(sheename_rownum, dialogue)
                data_list = []

                for idx, row in enumerate(sheet.iter_rows(min_row=8), start=8):  # idx从8开始
                    identifier = f"{sheet.title}_{idx}"  # 这里增加sheet名和行号
                    role = row[0].value if row[0].value else ""
                    dialogue = row[1].value if row[1].value else ""  # None 替换为空字符串

                    if role or dialogue:  # 如果角色或对话至少有一个非空
                        if data_list and not role:  # 如果角色为空，但数据列表不为空，表示这是前一个角色的继续
                            data_list[-1][1] += " " + dialogue
                        else:
                            data_list.append([identifier, role + ':' + dialogue])

                data.extend(data_list)

            # 创建新的工作簿和工作表
            workbook = Workbook()
            sheet = workbook.active

            # 设置headers
            headers = ["标识名*", "文本描述*", "正向提示词", "反向提示词", "状态"]
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

            # 设置内容
            for row_num, (identifier, description) in enumerate(data, start=2):
                sheet.cell(row=row_num, column=1, value=identifier)
                sheet.cell(row=row_num, column=2, value=description)

            # 保存文件
            workbook.save(self.role_file)
            self.add_string_to_listView("文本描述保存成功")
            self.SuggestKeywordsProgressBar.setValue(100)
        except:
            self.add_string_to_listView("文本描述保存失败，请检查“场景人物生成.xlsx”是否被占用")

    def viewfile(self):
        # 获取当前工作目录的路径
        project_root = os.getcwd()
        # 使用os.startfile打开当前项目的根目录
        os.startfile(project_root)

    def script(self):
        # 获取脚本表格目录名称
        dir_name = os.path.dirname(self.open_filepath)
        # 在该目录中设置新的文件名，即预处理后的文件
        self.solve_script_path = os.path.join(dir_name, "solve_script.xlsx")
        try:
            # 进行预处理
            script_preprocess(self.open_filepath, self.solve_script_path)
            # 更行预览框的各种字符串
            self.SheetComboBox.currentIndexChanged[str].disconnect(self.changeSheet)
            self.add_string_to_listView("打开文件路径" + self.solve_script_path)
            print(self.solve_script_path)
            # 加载Excel文件中的所有sheet名到ComboBox中
            self.workbook = load_workbook(self.solve_script_path)
            self.SheetComboBox.clear()
            self.SheetComboBox.addItems(self.workbook.sheetnames)
            # 默认加载第一个sheet的数据
            self.loadExcelData(self.workbook.sheetnames[0])
            self.SheetComboBox.currentIndexChanged[str].connect(self.changeSheet)
            self.add_string_to_listView("脚本预处理完成，输出文件“solve_script.xlsx”")
            try:
                # 多线程处理
                self.thread = ScriptThread(self.solve_script_path)
                self.thread.string_signal.connect(self.add_string_to_listView)
                self.thread.finished.connect(self.thread.deleteLater)
                self.thread.start()
            except Exception as e:
                self.add_string_to_listView("脚本生成多线程处理失败，请重试")
                print("脚本生成多线程处理失败", e)
        except Exception as e:
            self.add_string_to_listView("脚本预处理失败，请检查”solve_script.xlsx”是否被占用")
            print("脚本预处理失败", e)



