# -*- coding: utf-8 -*-
import json
import os

# Form implementation generated from reading ui file 'vits_window.ui'
#
# Created by: PyQt5 UI code generator 5.15.9
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QStringListModel, QThread, pyqtSignal
from datetime import datetime
from PyQt5.QtWidgets import QHeaderView, QFileDialog, QTableWidgetItem, QComboBox
from openpyxl.reader.excel import load_workbook
import re
from vits_vits import vits

# vits线程类
class VitsThread(QThread):
    # 更新进度的信号量
    progress_signal = pyqtSignal(int)
    # 更新日志的信号量
    string_signal = pyqtSignal(str)

    def __init__(self, data, sheetname, language, speed):
        super(VitsThread, self).__init__()
        self.data = data
        self.sheetname = sheetname
        self.language = language
        self.speed = speed
        self.running = True

    def run(self):
        i = 0
        for key, inner_dict in self.data.items():
            if self.running:
                character = inner_dict['value0']
                message = inner_dict['value1']
                people = inner_dict['value2']
                vits(message,people,self.language,self.speed,key,self.sheetname,character)
                self.string_signal.emit("audio/"+str(self.sheetname)+"_"+str(character)+"_"+str(key)+".wav保存完成")
                self.progress_signal.emit(i + 1)
                i += 1
        self.string_signal.emit("语音合成成功")

class Ui_Vits(object):
    def setupUi(self, Vits):
        Vits.setObjectName("Vits")
        Vits.resize(1711, 1041)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        Vits.setPalette(palette)
        # self.centralwidget = QtWidgets.QWidget(Vits)
        # self.centralwidget.setObjectName("centralwidget")
        self.VitsListView = QtWidgets.QListView(Vits)
        self.VitsListView.setGeometry(QtCore.QRect(1350, 11, 341, 241))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(24, 24, 24))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.VitsListView.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.VitsListView.setFont(font)
        self.VitsListView.setStyleSheet("background-color: rgb(24, 24, 24); color: rgb(255, 255, 255);")
        self.VitsListView.setObjectName("VitsListView")
        self.tableWidget = QtWidgets.QTableWidget(Vits)
        self.tableWidget.setGeometry(QtCore.QRect(40, 270, 1651, 731))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.tableWidget.sizePolicy().hasHeightForWidth())
        self.tableWidget.setSizePolicy(sizePolicy)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.layoutWidget = QtWidgets.QWidget(Vits)
        self.layoutWidget.setGeometry(QtCore.QRect(20, 10, 1331, 47))
        self.layoutWidget.setObjectName("layoutWidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.layoutWidget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.ImportConfigurationButton = QtWidgets.QPushButton(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.ImportConfigurationButton.sizePolicy().hasHeightForWidth())
        self.ImportConfigurationButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.ImportConfigurationButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.ImportConfigurationButton.setFont(font)
        self.ImportConfigurationButton.setStyleSheet("background-color: rgb(0, 104, 0);")
        self.ImportConfigurationButton.setObjectName("ImportConfigurationButton")
        self.horizontalLayout.addWidget(self.ImportConfigurationButton)
        self.StartVitsButton = QtWidgets.QPushButton(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.StartVitsButton.sizePolicy().hasHeightForWidth())
        self.StartVitsButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.StartVitsButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.StartVitsButton.setFont(font)
        self.StartVitsButton.setStyleSheet("background-color: rgb(0, 104, 0);")
        self.StartVitsButton.setObjectName("StartVitsButton")
        self.horizontalLayout.addWidget(self.StartVitsButton)
        # self.StartVitsButton_2 = QtWidgets.QPushButton(self.layoutWidget)
        # sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        # sizePolicy.setHorizontalStretch(0)
        # sizePolicy.setVerticalStretch(0)
        # sizePolicy.setHeightForWidth(self.StartVitsButton_2.sizePolicy().hasHeightForWidth())
        # self.StartVitsButton_2.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        # self.StartVitsButton_2.setPalette(palette)
        # font = QtGui.QFont()
        # font.setFamily("阿里巴巴普惠体")
        # font.setPointSize(20)
        # font.setBold(True)
        # font.setWeight(75)
        # self.StartVitsButton_2.setFont(font)
        # self.StartVitsButton_2.setStyleSheet("background-color: rgb(0, 104, 0);")
        # self.StartVitsButton_2.setObjectName("StartVitsButton_2")
        # self.horizontalLayout.addWidget(self.StartVitsButton_2)
        self.PauseButton = QtWidgets.QPushButton(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.PauseButton.sizePolicy().hasHeightForWidth())
        self.PauseButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(222, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.PauseButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.PauseButton.setFont(font)
        self.PauseButton.setStyleSheet("background-color: rgb(222, 0, 0);")
        self.PauseButton.setObjectName("PauseButton")
        self.horizontalLayout.addWidget(self.PauseButton)
        self.ViewVitsFileButton = QtWidgets.QPushButton(self.layoutWidget)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.ViewVitsFileButton.sizePolicy().hasHeightForWidth())
        self.ViewVitsFileButton.setSizePolicy(sizePolicy)
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(120, 120, 120))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(0, 104, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        self.ViewVitsFileButton.setPalette(palette)
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.ViewVitsFileButton.setFont(font)
        self.ViewVitsFileButton.setStyleSheet("background-color: rgb(0, 104, 0);")
        self.ViewVitsFileButton.setObjectName("ViewVitsFileButton")
        self.horizontalLayout.addWidget(self.ViewVitsFileButton)
        self.SheetComboBox = QtWidgets.QComboBox(Vits)
        self.SheetComboBox.setGeometry(QtCore.QRect(40, 180, 161, 51))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.SheetComboBox.setFont(font)
        self.SheetComboBox.setStyleSheet("border: 2px solid grey;\n"
"border-radius: 5px;")
        self.SheetComboBox.setObjectName("SheetComboBox")
        self.label_2 = QtWidgets.QLabel(Vits)
        self.label_2.setGeometry(QtCore.QRect(40, 110, 111, 37))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(Vits)
        self.label_3.setGeometry(QtCore.QRect(230, 110, 111, 37))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.PeopleComboBox = QtWidgets.QComboBox(Vits)
        self.PeopleComboBox.setGeometry(QtCore.QRect(230, 180, 161, 51))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.PeopleComboBox.setFont(font)
        self.PeopleComboBox.setStyleSheet("border: 2px solid grey;\n"
"border-radius: 5px;")
        self.PeopleComboBox.setObjectName("PeopleComboBox")
        self.label_4 = QtWidgets.QLabel(Vits)
        self.label_4.setGeometry(QtCore.QRect(420, 110, 111, 37))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.LanguageComboBox = QtWidgets.QComboBox(Vits)
        self.LanguageComboBox.setGeometry(QtCore.QRect(420, 180, 161, 51))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.LanguageComboBox.setFont(font)
        self.LanguageComboBox.setStyleSheet("border: 2px solid grey;\n"
"border-radius: 5px;")
        self.LanguageComboBox.setObjectName("LanguageComboBox")
        self.label_5 = QtWidgets.QLabel(Vits)
        self.label_5.setGeometry(QtCore.QRect(610, 110, 111, 37))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.SpeedSlider = QtWidgets.QSlider(Vits)
        self.SpeedSlider.setGeometry(QtCore.QRect(610, 190, 111, 22))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.SpeedSlider.setFont(font)
        self.SpeedSlider.setMaximum(5)
        self.SpeedSlider.setSingleStep(1)
        self.SpeedSlider.setProperty("value", 1)
        self.SpeedSlider.setOrientation(QtCore.Qt.Horizontal)
        self.SpeedSlider.setInvertedAppearance(False)
        self.SpeedSlider.setInvertedControls(False)
        self.SpeedSlider.setTickPosition(QtWidgets.QSlider.NoTicks)
        self.SpeedSlider.setObjectName("SpeedSlider")
        self.label = QtWidgets.QLabel(Vits)
        self.label.setGeometry(QtCore.QRect(851, 110, 211, 37))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")

        self.tableWidget.setFont(font)

        self.VitsProgressBar = QtWidgets.QProgressBar(Vits)
        self.VitsProgressBar.setEnabled(True)
        self.VitsProgressBar.setGeometry(QtCore.QRect(850, 180, 471, 49))
        font = QtGui.QFont()
        font.setFamily("阿里巴巴普惠体")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.VitsProgressBar.setFont(font)
        self.VitsProgressBar.setStyleSheet("QProgressBar {\n"
"   border: 2px solid grey;\n"
"   border-radius: 5px;\n"
"   background-color: #FFFFFF;\n"
"}\n"
"\n"
"QProgressBar::chunk {\n"
"   background-color: #0000FF;\n"
"   width: 20px;\n"
"}\n"
"\n"
"QProgressBar {\n"
"   border: 2px solid grey;\n"
"   border-radius: 5px;\n"
"   text-align: center;\n"
"}")
        self.VitsProgressBar.setProperty("value", 0)
        self.VitsProgressBar.setObjectName("VitsProgressBar")
        # Vits.setCentralWidget(self.centralwidget)
        # self.menubar = QtWidgets.QMenuBar(Vits)
        # self.menubar.setGeometry(QtCore.QRect(0, 0, 1711, 23))
        # self.menubar.setObjectName("menubar")
        # Vits.setMenuBar(self.menubar)
        # self.statusbar = QtWidgets.QStatusBar(Vits)
        # self.statusbar.setObjectName("statusbar")
        # Vits.setStatusBar(self.statusbar)


        self.retranslateUi(Vits)
        QtCore.QMetaObject.connectSlotsByName(Vits)

        # 获得speakers_list
        self.speakers_list = []
        self.get_speakers_list()

        self.log_data = []
        # 设置配音下拉列表
        self.PeopleComboBox.addItems(self.speakers_list)
        # 设置语种下拉列表
        self.LanguageComboBox.addItems(["日本語", "简体中文", "English", "Mix", ])
        # 绑定button到槽函数
        self.ImportConfigurationButton.clicked.connect(self.open_file)
        self.SheetComboBox.currentTextChanged.connect(self.loadSelectedSheet)
        self.PeopleComboBox.currentTextChanged.connect(self.loadPeople)
        self.StartVitsButton.clicked.connect(self.startVitsThread)
        self.PauseButton.clicked.connect(self.stop)
        self.ViewVitsFileButton.clicked.connect(self.viewfile)


    def retranslateUi(self, Vits):
        _translate = QtCore.QCoreApplication.translate
        Vits.setWindowTitle(_translate("Vits", "MainWindow"))
        self.ImportConfigurationButton.setText(_translate("Vits", "从Excel文件导入配置"))
        self.StartVitsButton.setText(_translate("Vits", "开始合成语音（当前表单）"))
        #self.StartVitsButton_2.setText(_translate("Vits", "开始合成语音（所有表单）"))
        self.PauseButton.setText(_translate("Vits", "一键暂停"))
        self.ViewVitsFileButton.setText(_translate("Vits", "查看文件"))
        self.label_2.setText(_translate("Vits", "选择表单"))
        self.label_3.setText(_translate("Vits", "默认配音"))
        self.label_4.setText(_translate("Vits", "配音语种"))
        self.label_5.setText(_translate("Vits", "配音语速"))
        self.label.setText(_translate("Vits", "Vits语音合成进度"))

    def set_listView_content(self):
        # 使用QStringListModel来设置listView的数据
        model = QStringListModel()
        model.setStringList(self.log_data)
        self.VitsListView.setModel(model)

    def add_string_to_listView(self, new_string):
        # 获取当前时间
        current_time = datetime.now()
        # 将时间转换为字符串格式
        time_string = current_time.strftime('%m-%d %H:%M:%S ')
        new_string = time_string + new_string
        self.log_data.append(new_string)
        # 更新 listView 的内容
        self.set_listView_content()

    def open_file(self):
        self.open_filepath, _ = QFileDialog.getOpenFileName(None, '选择Excel配置文件(路径尽量不要有中文)', '',
                                                            '(*.xlsx *.xls)')
        try:
            # 替换为单反斜杠
            self.open_filepath = str(self.open_filepath).replace("/", "\\").replace(":", ":")
            self.open_filename = os.path.basename(self.open_filepath)
            self.add_string_to_listView("打开文件路径" + self.open_filepath)
            print(self.open_filepath)
            self.loadExcelData()
        except:
            self.add_string_to_listView("取消打开文件")
            print("取消打开文件")

    def loadExcelData(self):
        # 清空tableWidget的内容
        self.tableWidget.setRowCount(0)
        self.tableWidget.setColumnCount(0)
        # 使用openpyxl加载Excel文件数据
        self.workbook = load_workbook(self.open_filepath)
        self.active_sheet = self.workbook.active

        # 加载Excel文件并且将所有工作表加入QComboBox
        self.SheetComboBox.clear()
        self.SheetComboBox.addItems(self.workbook.sheetnames)


    def loadSelectedSheet(self, sheet_name):
        try:
            self.tableWidget.setRowCount(0)
            self.active_sheet = self.workbook[sheet_name]

            # 在table widget中设置行列
            self.tableWidget.setRowCount(self.active_sheet.max_row)
            min_column = 1
            max_column = 3
            min_row = 8
            self.tableWidget.setColumnCount(max_column - min_column + 1)

            # 使用Excel的第一行作为table widget的头
            # 设置静态的表头为 ["角色", "话", "配音人员"]
            headers = ["角色", "话", "配音人员"]
            self.tableWidget.setHorizontalHeaderLabels(headers)

            # 临时列表存储有效行
            valid_rows = []

            # 载入每条数据
            for row in self.active_sheet.iter_rows(min_row=min_row, max_col=max_column):
                col1_value = row[0].value if row[0].value is not None else ""
                col2_value = row[1].value if len(row) > 1 and row[1].value is not None else ""

                if col1_value or col2_value:  # 检查至少有一个是有数据的，则算有效行
                    valid_rows.append(row)

            # 根据有效行设置tableWidget
            self.tableWidget.setRowCount(len(valid_rows))

            unique_values = set()  # 创建一个空的集合来存储唯一的值

            # 载入数据
            for idx, valid_row in enumerate(valid_rows):
                for cell in valid_row:
                    value = cell.value if cell.value is not None else ""
                    self.tableWidget.setItem(idx, cell.column - min_column, QTableWidgetItem(str(value)))
                    # 检查第0列的值是否为空、是否重复,不为则加入配音人员选择
                    col0_value = valid_row[0].value if valid_row[0].value is not None else ""
                    if col0_value and col0_value not in unique_values:
                        self.set_tableWidget_combobox(idx, 2)
                        unique_values.add(value)  # 向集合中添加值，如果该值已存在，则不会重复添加

            # 设置行高度
            for rowIndex in range(self.tableWidget.rowCount()):
                self.tableWidget.setRowHeight(rowIndex, 200)
            # 列自适应，(水平方向占满窗体)
            self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            # 行自适应，(垂直方向占满窗体)
            # self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        except:
            print()

    # 获取speakers_list
    def get_speakers_list(self):
        # 1. 打开并读取vits.json文件
        with open('vits.json', 'r', encoding='utf-8') as file:
            data = json.load(file)
        # 2. 获取speakers字段
        speakers_dict = data.get("speakers", {})
        # 3. 将字典的键提取到一个列表中
        self.speakers_list = list(speakers_dict.keys())
        print(self.speakers_list)

    def set_tableWidget_combobox(self, row, column):
        combobox = QComboBox()  # 创建一个QComboBox对象

        # 为combobox添加一些选项：
        combobox.addItems(self.speakers_list)
        combobox.setStyleSheet("""
        QComboBox { 
            background-color: lightblue;
            text-align: center;
            font: 20pt "阿里巴巴普惠体"
        }
        """)
        self.tableWidget.setCellWidget(row, column, combobox)

    def loadPeople(self):
        unique_values = set()  # 创建一个空的集合来存储唯一的值

        for row in range(self.tableWidget.rowCount()):
            # 获取第0列的QTableWidgetItem
            item = self.tableWidget.item(row, 0)
            # 检查第0列的值是否为空
            value = item.text() if item.text() is not None else ""
            if value and value not in unique_values:
                combobox = QComboBox()  # 创建一个QComboBox对象
                # 为combobox添加一些选项：
                combobox.addItems(self.speakers_list)
                # 设置self.PeopleComboBox.currentText()为当前选中值
                combobox.setCurrentText(self.PeopleComboBox.currentText())
                combobox.setStyleSheet("""
QComboBox { 
    background-color: lightblue;
    text-align: center;
    font: 20pt "阿里巴巴普惠体"
}
""")
                self.tableWidget.setCellWidget(row, 2, combobox)
                unique_values.add(value)  # 向集合中添加值，如果该值已存在，则不会重复添加

    # 按顺序，从第0行开始，将0作为key，第0列的值作为value0，第1列作为value1，第2列作为value2。
    # 此外，第0列的值可能为空，所以如果第i行0列的值为空，则将i行0列的值为上一次不为空的值（比如上一次不为空的值为i-2行0列，key=i-2，value0=a，value1=b，value2=c，则第i行0列的key=i，value0=a，value1=b，value2=c）
    def getTableData(self):
        result = {}
        result_by_value0 = {}
        last_non_empty_value0 = None
        last_non_empty_value2 = None

        for row in range(self.tableWidget.rowCount()):
            # 获取第0列的值
            value0 = self.tableWidget.item(row, 0)
            value0_text = value0.text() if value0 and value0.text().strip() != "" else None

            # 获取第2列的QComboBox选中的字符串
            combobox = self.tableWidget.cellWidget(row, 2)
            if combobox and isinstance(combobox, QComboBox):
                value2_text = combobox.currentText()
            else:
                value2_text = ""

            # 如果第0列的值为空，使用上一个非空值；否则，使用当前的值并更新last_non_empty_value
            if value0_text:
                last_non_empty_value0 = value0_text
            else:
                value0_text = last_non_empty_value0

            # 获取第1列的值
            value1 = self.tableWidget.item(row, 1)
            value1_text = value1.text() if value1 else ""

            if value2_text != "":
                result_by_value0[value0_text] = value2_text

            if (not value1_text.startswith("跳转到Sheet")):
                # 如果话这一列的值不是“跳转到Sheet”，则将数据存储在字典中
                # 正则表达式 [^\u4e00-\u9fa5，。！] 代表所有不是中文字符、逗号和句号的字符，然后我们使用 re.sub() 方法替换这些字符为“”。
                value1_text = re.sub(r"[^\u4e00-\u9fa5，。！、—？]", "", value1_text)
                result[row] = {"value0": value0_text, "value1": value1_text, "value2": result_by_value0[value0_text]}

        return result

    def startVitsThread(self):
        self.add_string_to_listView("开始合成语音...")
        data = self.getTableData()
        self.VitsProgressBar.setValue(0)  # 设置进度条的最小值
        self.VitsProgressBar.setMaximum(int(len(data)))  # 设置进度条的最大值
        self.thread = VitsThread(data, self.SheetComboBox.currentText(), self.LanguageComboBox.currentText(), self.SpeedSlider.value())
        self.thread.progress_signal.connect(self.VitsProgressBar.setValue)
        self.thread.string_signal.connect(self.add_string_to_listView)
        self.thread.finished.connect(self.thread.deleteLater)
        self.thread.start()

    def stop(self):
        self.thread.running = False
        self.add_string_to_listView("所有操作已暂停")

    def viewfile(self):
        os.startfile('audio')